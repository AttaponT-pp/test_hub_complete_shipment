import sys
import re
from os import path
import datetime, time

import xlrd
from openpyxl import *
from PyQt5 import QtWidgets
from win32api import Unicode
import Form

import ctypes
from FITS_Connect import *


class CompleteShipmentMainUi(QtWidgets.QMainWindow):

    def __init__(self):
        super(CompleteShipmentMainUi, self).__init__()
        self.form = Form.Ui_MainForm()
        self.form.setupUi(self)
        self.setFixedSize(380, 340)
        self.form.tabWidget.setCurrentIndex(0)
        self.fill_en()
        # connect events
        self.form.txten.returnPressed.connect(self.page_focus)
        self.form.tabWidget.currentChanged.connect(self.tab_on_change)
        self.form.btn_oba_browse.clicked.connect(self.select_oba_file)
        self.form.btn_rtv_browse.clicked.connect(self.select_rtv_file)
        self.form.txt_inv.textChanged.connect(self.inv_count)
        self.form.txt_etr.textChanged.connect(self.etr_count)
        self.form.txt_inv.returnPressed.connect(self.get_invoice)
        self.form.txt_etr.returnPressed.connect(self.get_etr)

    def tab_on_change(self):
        tab_index = self.form.tabWidget.currentIndex()
        print(self.form.tabWidget.currentIndex())
        print(len(self.form.txten.text()))
        if len(self.form.txten.text()) != 6:
            self.fill_en()
        else:
            if tab_index == 0:
                if self.form.oba_file_path.toPlainText() == '':
                    self.form.btn_oba_browse.setFocus()
                    self.form.lbl_app_status.setText('Browse your OBA Request File.')
                else:
                    self.form.txt_inv.setFocus()
                    self.form.txt_inv.setText('')
                    self.form.lbl_app_status.setText('Scan Invoice#')

            else:
                if self.form.rtv_file_path.toPlainText() == '':
                    self.form.btn_rtv_browse.setFocus()
                    self.form.lbl_app_status.setText('Browse your Daily Shipment Request File.')
                else:
                    self.form.txt_etr.setFocus()
                    self.form.txt_etr.setText('')
                    self.form.lbl_app_status.setText('Scan ETR No.')

    def get_invoice(self):
        now = datetime.datetime.now()
        self.form.textEdit.setStyleSheet("background-color: rgb(0, 0, 0);")
        f_path = str(self.form.oba_file_path.toPlainText())
        filled_en = str(self.form.txten.text())
        f_data = {'rt': '', 'packing_lot': '', 'po_num': '', 'part_num': '', 'qty': '', 'inv': ''}

        # Check EN
        if filled_en == "" or not len(filled_en) == 6:
            mbox(u'Please Fill EN Again', u'Please Fill EN Again', 0)
            self.form.txt_etr.setText("")
            self.fill_en()
            return
        # Check file path
        if f_path == "":
            mbox(u'Not Found OBA Request File', u'Please Browse OBA Request before Enter ETR', 0)
            self.form.txt_inv.setText("")
            return
        # Check file exist
        if not path.isfile(f_path):
            mbox(u'Not Found OBA Request File', u'File Not Found! Please browse the file again', 0)
            self.form.txt_inv.setText("")
            self.form.oba_file_path.setText("")
            return

        inv_num = str(self.form.txt_inv.text().strip())

        # check length of Invoice#
        if len(inv_num) != 7:
            mbox('Wrong invoice number format', "Please re-scan invoice number", 0)
            self.form.txt_inv.setText('')
            return

        # check invoice# in OBA Request
        if not cross_check_inv(f_path, inv_num, f_data):
            mbox(u'Not Found Invoice number', u'Not Found Invoice# in OBA Request', 0)
            self.form.txt_inv.setText('')
            return

        if f_data['inv'] != '':
            tmp_inv = f_data['inv']
            print(tmp_inv)
            self.form.lbl_app_status.setText('Found Invoice# {}'.format(tmp_inv))
            # get check box operation
            which_opn = self.check_opn_box()
            if not which_opn['opn1501'] and not which_opn['opn702']:
                mbox('No FITS OPN select.', 'Not found FITS OPN selected.\nPlease select target OPN.', 0)
                return
            if which_opn['opn1501']:
                # Opn.1501 OBA
                # check packing_lot
                if f_data['packing_lot'] == 'None':
                    print('Packing Number is empty.')
                    # get Packing Number from FITS
                    f_data['packing_lot'] = find_packing_num(f_data['rt'])

                # prepare data_stream
                packing_no = f_data['packing_lot']
                oba_info = prepare_oba_info(packing_no)
                print(oba_info)

                # create data1501
                data1501 = filled_en + ',' + inv_num + ',' + f_data['packing_lot'] + ',' + oba_info
                print('Opn.1501 param: {}'.format(opn1501_param))
                print('Data stream: {}'.format(data1501))
                # hand_check data and operation
                inv_last_opn = get_last_opn(inv_num)
                print(inv_last_opn)
                route_check = valid_inv('1501', inv_num)
                print('Hand-Check: {}'.format(route_check['status']))
                if route_check['status']:
                    # save FITS data
                    print('Recording data...')
                    if record2fit('1501', opn1501_param, data1501):
                        print('Save data opn.1501 OBA completed.')
                        self.form.txt_inv.setFocus()
                        self.form.txt_inv.setText('')
                        self.form.textEdit.setStyleSheet("background-color: rgb(0, 255, 0);")
                        self.form.lbl_app_status.setText('Invoice no.{} is saved successful.'.format(inv_num))
                    else:
                        print('Cannot save data opn.1501 OBA.')
                        self.form.txt_inv.setFocus()
                        self.form.txt_inv.setText('')
                        self.form.textEdit.setStyleSheet("background-color: rgb(255, 0, 0);")
                        self.form.lbl_app_status.setText('FITS Error: Cannot save opn.1501')
                        return
                else:
                    self.form.lbl_app_status.setText('FITSDLL Error: {}'.format(route_check["msg"]))
                    self.form.textEdit.setStyleSheet("background-color: rgb(255, 0, 0);")
                    mbox(u'FITSDLL Error', route_check["msg"], 0)
                    return

            if which_opn['opn702']:
                # Opn.702 Shipment SN
                sn_list = get_sn_list(f_data['rt'])
                print(sn_list)
                for sn in sn_list.split(','):
                    print('Serial No.:{}'.format(sn))
                    print('Validate route...')
                    route_check = valid_inv('702', sn)
                    print(route_check['status'])
                    if route_check['status']:
                        # get last operation
                        print('Get last operation...')
                        last_opn = get_last_opn(sn)
                        print('Last operation of SN: {} is {}'.format(sn, last_opn))
                        if last_opn == '601_B':
                            print('Ready to save...')
                            # prepare data_stream
                            data702 = filled_en + ',' + inv_num + ',' + sn + ',' + f_data['packing_lot'] + ',' + str(f_data['qty'])
                            print('Data stream: {}'.format(data702))
                            print('Opn.702 param: {}'.format(opn702_param))
                            print('Record data...')
                            if record2fit('702', opn702_param, data702):
                                print('Save SN: {} successful.'.format(sn))
                                self.form.txt_inv.setFocus()
                                self.form.txt_inv.setText('')
                                self.form.textEdit_2.setStyleSheet("background-color: rgb(0, 255, 0);")
                                self.form.lbl_app_status.setText('Invoice no.{} : {} successful.'.format(inv_num, sn))
                            else:
                                print('Save SN: {} error.'.format(sn))
                                self.form.txt_inv.setFocus()
                                self.form.txt_inv.setText('')
                                self.form.textEdit_2.setStyleSheet("background-color: rgb(255, 0, 0);")
                                self.form.lbl_app_status.setText('FITS Error: Cannot save opn.702')
                                return
                        else:
                            self.form.textEdit_2.setStyleSheet("background-color: rgb(255, 0, 0);")
                            self.form.lbl_app_status.setText('This SN: {} does not pack yet'.format(sn))
                            print('This SN: {} does not pack yet'.format(sn))
                    else:
                        print('SN:{} route check error.'.format(sn))
                        self.form.lbl_app_status.setText('SN: {} route check error.'.format(sn))
                        self.form.textEdit_2.setStyleSheet("background-color: rgb(255, 0, 0);")
        else:
            mbox('Not found Invoice# {}'.format(inv_num), 'Not found Invoice# {} in OBA Summary File'.format(inv_num), 0)
            return

    def get_etr(self):
        now = datetime.datetime.now()
        f_path = str(self.form.rtv_file_path.toPlainText())
        filled_en = str(self.form.txten.text())
        f_data = {'inv': '', 'rt': '', 'qty': ''}

        # Check EN
        if filled_en == "" or not len(filled_en) == 6:
            mbox(u'Please Fill EN Again', u'Please Fill EN Again', 0)
            self.form.txt_etr.setText("")
            self.fill_en()
            return

        # Check file path
        if f_path == "":
            mbox(u'Not Found File Daily Shipment Request', u'Please Browse Daily Shipment Request before Enter ETR', 0)
            self.form.txt_etr.setText("")
            return

        # Check file exist
        if not path.isfile(f_path):
            mbox(u'Not Found File Daily Shipment Request', u'File Not Found! Please browse the file again', 0)
            self.form.txt_etr.setText("")
            self.form.rtv_file_path.setText("")
            return

        etr = str(self.form.txt_etr.text())

        # Check prefix of ETR
        if not etr[0].upper() == "C" and not etr[0].upper() == "R":
            mbox(u'ETR wrong format', u'Please Enter ETR in correct format', 0)
            self.form.txt_etr.setText("")
            return

        # Check across this year
        # if not etr[1:5] == str(now.year):
        # Mbox(u'ETR wrong format', u'Please Enter ETR in correct format', 0)
        # self.line_etr.setText("")
        # return

        # Check length of ETR
        if not len(etr) == 9:
            mbox(u'ETR wrong format', u'Please Enter ETR in correct format', 0)
            self.form.txt_etr.setText("")
            return

        # Check ETR across file
        if not cross_check_etr(f_path, etr, f_data):
            mbox(u'Not Found ETR', u'Not Found ETR in File Daily Shipment Request', 0)
            self.form.txt_etr.setText("")
            return

        # Print Invoice on Label
        # if f_data['inv'] != "None" and f_data['rt'] != "None" and f_data['qty'] != "None":
        if f_data['inv'] != "None":
            tmp_inv = f_data['inv']
            # tmp_rt = f_data['rt']
            # tmp_qty = f_data['qty']

            # Get RTV Shipment Blocking
            data1502 = str(prepare_etr_info(etr))
            block_rtv_status = data1502.split(',')[6]
            if block_rtv_status == "YES":
                mbox("Warning !!!", "This ETR Number:" + etr + ". have been blocked in Opn.924 RTV Shipment Blocking"
                                                               "Please inform case owner for unblock.")
                self.form.lbl_app_status.setText("ETR Number:" + etr + " have been blocked RTV Shipment")
                self.form.txt_etr.setFocus()
                self.form.txt_etr.setText("")
                return

            self.form.lbl_app_status.setText('Found Invoice {}'.format(tmp_inv))
            which_opn = self.check_opn_box()

            if not which_opn["opn1502"] and not which_opn["opn1801"]:
                mbox(u'No FITS OPN Selected', u'Please Select FITS Operation to record.', 0)
                return

            if which_opn["opn1502"]:
                # Create data_str input stream
                data_str = filled_en + ',' + tmp_inv + ',' + etr + ',' + block_rtv_status
                fit_status = valid_inv('1502', tmp_inv)
                if fit_status["status"]:
                    # Add Parameter 'RTV Shipment Blocking'
                    param1502_param = 'OPERATOR,Invoice No,ETR Number,Part Number,Supplier Name,RT,PO No.,' \
                                      'RTV Shipment Request,Build Type,RTV Shipment Blocking'
                    param1502_str = 'OPERATOR,Invoice No,ETR Number,RTV Shipment Blocking'
                    print(param1502_str)
                    print(param1502_param)
                    print(data_str)
                    # if record2fit('1502', param1502_param, data_str):
                    if record2fit('1502', param1502_str, data_str):
                        print('ETR No.{} is save with Invoice No.{} successful.'.format(etr, tmp_inv))
                        self.form.lbl_app_status.setText('FITS1502 Saved for {}'.format(etr))
                        self.form.txt_etr.setText("")
                        self.form.txt_etr.setFocus()
                        self.form.textEdit_3.setStyleSheet("background-color: rgb(0, 255, 0);")
                    else:
                        mbox(u'FITSDLL Error', u'Cannot save data to FITS', 0)
                        self.form.textEdit_3.setStyleSheet("background-color: rgb(255, 0, 0);")
                        return
                else:
                    mbox(u'FITSDLL Error', Unicode(fit_status["msg"]), 0)
                    self.form.textEdit_3.setStyleSheet("background-color: rgb(255, 0, 0);")
                    return

            if which_opn["opn1801"]:
                time.sleep(1)
                init('1801')
                # Create data_str input stream
                data_str = filled_en + ',' + tmp_inv + ',' + etr
                data1303 = get_necessory_data('1303', etr, 'Serial No,Fail Qty')
                print(data1303)

                # fit_status = valid_inv('1801', data1303[0])
                if init('1801') == 'True':
                    param1801_str = 'OPERATOR,Invoice No,ETR Number,Serial No,ETR Lot Qty'
                    if len(data1303) < 3:
                        ship_data_str = data_str + ',' + data1303[0] + ',' + data1303[1]
                        print(ship_data_str)

                        if record2fit('1801', param1801_str, ship_data_str) == 'True':
                            self.form.lbl_app_status.setText('FITS1801 Saved for {}'.format(etr))
                            self.form.txt_etr.setText("")
                            self.form.txt_etr.setFocus()
                            self.form.textEdit_4.setStyleSheet("background-color: rgb(0, 255, 0);")
                        else:
                            mbox(u'FITSDLL Error', u'Cannot save data to FITS', 0)
                            self.form.textEdit_4.setStyleSheet("background-color: rgb(255, 0, 0);")
                            return
                    else:
                        for sn in data1303:
                            ship_data_str = ""
                            if len(sn) == 11:
                                ship_data_str = data_str + ',' + sn + ',' + data1303[len(data1303) - 1]
                                print(ship_data_str)

                                if record2fit('1801', param1801_str, ship_data_str) == 'True':
                                    self.form.lbl_app_status.setText('FITS1801 Saved for {}'.format(etr))
                                    self.form.txt_etr.setText("")
                                    self.form.txt_etr.setFocus()
                                    self.form.textEdit_4.setStyleSheet("background-color: rgb(0, 255, 0);")
                                else:
                                    mbox(u'FITSDLL Error', u'Cannot save data to FITS', 0)
                                    self.form.textEdit_4.setStyleSheet("background-color: rgb(255, 0, 0);")
                                    return

                else:
                    mbox(u'FITSDLL Error', u'Unable to init FITs DB', 0)
                    self.form.textEdit_4.setStyleSheet("background-color: rgb(255, 0, 0);")
                    return

        else:
            mbox(u'Not Found Invoice/RT/Qty', u'Not Found Invoice/RT/Qty for ETR: {}'.format(etr), 0)
            return

            # 2nd get neccessory data from rt
            # param = 'Part No.,Supplier Name,Build Type,PO No.'
            # param1502_out = param.split(',')
            # print tmp_inv+',' +tmp_rt+','+ tmp_qty
            # data1502 = get_necessory_data('101',tmp_rt,param)
            # Valid output
            # for i in range(len(data1502)):

            # print tmp_param[i] + '=' + q_data[i] + '\n'
            # if data1502[i] == "-":
            # Mbox(u'No Data Found!', u'There is no data for {} from FIT.'.format(param1502_out[i]), 0)
            # self.CleanupCtrs()
            # return

            # f_data['qty'] = str(get_necessory_data('1303',etr,'Fail Qty')[0])

    def fill_en(self):
        self.form.lbl_app_status.setText('Please fill your EN')
        self.form.txten.setText('')
        self.form.txten.setFocus()
        return

    def page_focus(self):
        en = self.form.txten.text()
        oba_path = str(self.form.oba_file_path.toPlainText())
        rtv_path = str(self.form.rtv_file_path.toPlainText())
        if not len(en) == 6:
            mbox("EN Validation", "Your EN is not valid.\nPlease try again.", 0)
            self.fill_en()
            return
        else:
            if self.form.tabWidget.currentIndex() == 0:
                print('Page 0')
                if oba_path != '':
                    self.form.txt_inv.setFocus()
                    self.form.lbl_app_status.setText('Scan Invoice#')
                    return
                else:
                    self.form.btn_oba_browse.setFocus()
                    self.form.lbl_app_status.setText('Browse your OBA Request File.')
                    return
            else:
                print('Page 1')
                if rtv_path != '':
                    self.form.txt_etr.setFocus()
                    self.form.lbl_app_status.setText('Scan ETR No.')
                    return
                else:
                    self.form.btn_rtv_browse.setFocus()
                    self.form.lbl_app_status.setText('Browse your Daily Shipment Request File.')
                    return

    def select_oba_file(self):
        dlg = QtWidgets.QFileDialog.getOpenFileName()
        print(dlg)
        if dlg[0] == '':
            mbox('OBA Request File Selection', 'No OBA Request file select', 0)
            self.form.oba_file_path.setText(dlg[0])
        else:
            print('Selected file: {}'.format(dlg))
            self.form.oba_file_path.setText(dlg[0])
            self.form.txt_inv.setFocus()
            self.form.txt_inv.setText('')
            self.form.lbl_app_status.setText('Scan Invoice#')
            return

    def select_rtv_file(self):
        dlg = QtWidgets.QFileDialog.getOpenFileName()
        print(dlg)
        if dlg[0] == '':
            mbox('Daily Shipment Request File Selection', 'No Daily Shipment Request file select', 0)
            self.form.rtv_file_path.setText(dlg[0])
        else:
            print('Selected file: {}'.format(dlg))
            self.form.rtv_file_path.setText(dlg[0])
            self.form.txt_etr.setFocus()
            self.form.txt_etr.setText('')
            self.form.lbl_app_status.setText('Scan ETR No.')
            return

    def etr_count(self):
        self.form.lbl_etr_count.setText("({})".format(len(self.form.txt_etr.text())))
        return

    def inv_count(self):
        self.form.lbl_inv_count.setText("({})".format(len(self.form.txt_inv.text())))

    def check_opn_box(self):
        return {"opn1501": self.form.checkBox1501.isChecked(), "opn702": self.form.checkBox702.isChecked(),
                "opn1502": self.form.checkBox1502.isChecked(), "opn1801": self.form.checkBox1801.isChecked()}


def mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)


# using openpyxl
def cross_check_inv(xlsx_path, inv, f_data):
    print('Cross Check Invoice number# {}'.format(inv))
    oba_wb = load_workbook(xlsx_path)
    output = oba_wb.active
    print(output.title)
    for row in range(5, output.max_row):
        # get invoice#
        if output.cell(row=row, column=10).value is None:
            break
        tmp_inv = output.cell(row=row, column=10).value
        print(tmp_inv)
        if re.search(inv, tmp_inv, re.IGNORECASE):
            f_data['rt'] = output.cell(row=row, column=2).value
            print(f_data['rt'])
            f_data['packing_lot'] = output.cell(row=row, column=3).value
            print(f_data['packing_lot'])
            f_data['po_num'] = output.cell(row=row, column=4).value
            print(f_data['po_num'])
            f_data['part_num'] = output.cell(row=row, column=5).value
            print(f_data['part_num'])
            f_data['qty'] = output.cell(row=row, column=6).value
            print(f_data['qty'])
            inv_num = "{}".format(output.cell(row=row, column=10).value)
            f_data['inv'] = inv_num.split('\n')[0]
            print(f_data['inv'])
            return True
    return False


def cross_check_etr(xls_path, etr, f_data):
    print('Cross Check ETR: {}'.format(etr))
    rtv_wb = load_workbook(xls_path)
    shipment = rtv_wb.active
    print(shipment.title)
    for row in range(5, shipment.max_row):
        if shipment.cell(row=row, column=23).value is None:
            break
        # get ETR#
        tmp_etr = shipment.cell(row=row, column=23).value
        print(tmp_etr)
        if re.search(etr, tmp_etr, re.IGNORECASE):
            f_data['inv'] = str(shipment.cell(row=row, column=22).value)
            return True
    return False
    # rtv_workbook = xlrd.open_workbook(xls_path)
    # print(rtv_workbook.get_sheets())
    # shipment = rtv_workbook.sheet_by_name('Shipment')
    # print(shipment.name)
    # for row in range(0, shipment.nrows):
    #
    #     hold = shipment.cell_value(rowx=row, colx=22)
    #     # hold = shipment.cell_value(rowx=row, colx=23)
    #     print(hold)
    #     if re.search(etr, hold, re.IGNORECASE):
    #         inv_num = "{}" .format(shipment.cell_value(row, 21))
    #         # inv_num = "{}".format(shipment.cell_value(row, 22))
    #         f_data['inv'] = inv_num.replace(" ", "")
    #         print(f_data['inv'])
    #         # f_data['rt'] = str(shipment.cell(row, 4).value).replace(" ", "")
    #         # f_data['qty'] = str(shipment.cell(row, 9).value).replace(" ", "")
    #         return True
    # return False


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    main_app = CompleteShipmentMainUi()
    main_app.show()
    sys.exit(app.exec_())